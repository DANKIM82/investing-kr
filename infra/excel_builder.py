#!/usr/bin/env python3
"""
excel_builder.py - 재무 모델 Excel (.xlsx) 생성

JSON context에서 multi-tab Excel 생성. Income Statement / Balance Sheet / Cash Flow / DCF 표준 구조.

CLI:
    python infra/excel_builder.py \\
        --context reports/.tmp/model_context.json \\
        --output reports/{TICKER}_model.xlsx
"""

import argparse
import json
import os
import sys


def build_excel(context, output_path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[error] openpyxl 미설치. pip install openpyxl", file=sys.stderr)
        return False
    
    wb = Workbook()
    company = context.get("company", {})
    
    # 스타일 정의
    header_font = Font(name="Malgun Gothic", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
    section_font = Font(name="Malgun Gothic", size=10, bold=True)
    body_font = Font(name="Malgun Gothic", size=10)
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    
    # ============================================================
    # Tab 1: Cover
    # ============================================================
    ws = wb.active
    ws.title = "Cover"
    ws["A1"] = f"{company.get('name', '')} ({company.get('ticker', '')}) — Financial Model"
    ws["A1"].font = Font(name="Malgun Gothic", size=18, bold=True)
    ws.merge_cells("A1:F1")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    info_rows = [
        ("Ticker", company.get("ticker", "")),
        ("Market", company.get("market", "")),
        ("Currency", company.get("currency", "")),
        ("Sector", company.get("sector", "")),
        ("Industry", company.get("industry", "")),
        ("Generated", context.get("date", "")),
        ("Firm", context.get("firm_name", "Personal Research")),
    ]
    for i, (k, v) in enumerate(info_rows, start=3):
        ws.cell(row=i, column=1, value=k).font = section_font
        ws.cell(row=i, column=2, value=v).font = body_font
    
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 30
    
    # ============================================================
    # Tab 2: Assumptions
    # ============================================================
    ws = wb.create_sheet("Assumptions")
    assumptions = context.get("assumptions", {})
    ws["A1"] = "Key Assumptions"
    ws["A1"].font = header_font
    ws["A1"].fill = header_fill
    
    rows = [
        ("WACC", assumptions.get("wacc", 0.085)),
        ("Terminal Growth Rate", assumptions.get("terminal_growth", 0.025)),
        ("Tax Rate", assumptions.get("tax_rate", 0.24)),
        ("Risk-free Rate", assumptions.get("risk_free_rate", 0.035)),
        ("Equity Risk Premium", assumptions.get("erp", 0.06)),
        ("Beta", assumptions.get("beta", 1.0)),
        ("Cost of Debt (pre-tax)", assumptions.get("cost_of_debt", 0.045)),
        ("Scenario", context.get("scenario", "Base")),
    ]
    for i, (k, v) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=k).font = section_font
        cell = ws.cell(row=i, column=2, value=v)
        cell.font = body_font
        if isinstance(v, float):
            cell.number_format = "0.00%" if "Rate" in k or "WACC" in k or "Premium" in k or "Growth" in k else "0.00"
    
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    
    # ============================================================
    # Tab 3: Income Statement
    # ============================================================
    ws = wb.create_sheet("Income Statement")
    is_data = context.get("income_statement", {})
    periods = is_data.get("periods", [])
    
    # 헤더
    ws.cell(row=1, column=1, value="Income Statement").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    for j, p in enumerate(periods, start=2):
        cell = ws.cell(row=1, column=j, value=p)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # 행
    line_items = [
        ("Revenue / 매출액", "revenue"),
        ("  YoY Growth", "revenue_growth"),
        ("Cost of Revenue / 매출원가", "cost_of_revenue"),
        ("Gross Profit / 매출총이익", "gross_profit"),
        ("  Gross Margin", "gross_margin"),
        ("SG&A / 판관비", "sga"),
        ("Operating Income / 영업이익", "operating_income"),
        ("  Operating Margin", "operating_margin"),
        ("EBITDA", "ebitda"),
        ("  EBITDA Margin", "ebitda_margin"),
        ("Interest Expense / 이자비용", "interest_expense"),
        ("Pretax Income / 세전이익", "pretax_income"),
        ("Tax Expense / 법인세", "tax_expense"),
        ("Net Income / 당기순이익", "net_income"),
        ("  Net Margin", "net_margin"),
        ("Diluted EPS", "diluted_eps"),
    ]
    
    for i, (label, key) in enumerate(line_items, start=2):
        cell = ws.cell(row=i, column=1, value=label)
        cell.font = section_font if not label.startswith("  ") else body_font
        if label.startswith("  "):
            cell.font = Font(name="Malgun Gothic", size=9, italic=True, color="666666")
        
        values = is_data.get(key, [])
        for j, v in enumerate(values, start=2):
            if j - 2 < len(values):
                cell = ws.cell(row=i, column=j, value=values[j - 2] if v is not None else "")
                cell.font = body_font
                if "Margin" in label or "Growth" in label:
                    cell.number_format = "0.0%"
                elif "EPS" in label:
                    cell.number_format = "#,##0.00"
                else:
                    cell.number_format = "#,##0"
    
    ws.column_dimensions["A"].width = 35
    for j in range(2, len(periods) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 13
    
    # ============================================================
    # Tab 4: Balance Sheet
    # ============================================================
    ws = wb.create_sheet("Balance Sheet")
    bs_data = context.get("balance_sheet", {})
    bs_periods = bs_data.get("periods", periods)
    
    ws.cell(row=1, column=1, value="Balance Sheet").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    for j, p in enumerate(bs_periods, start=2):
        cell = ws.cell(row=1, column=j, value=p)
        cell.font = header_font
        cell.fill = header_fill
    
    bs_items = [
        ("Cash & Equivalents / 현금", "cash_and_equivalents"),
        ("Trade Receivables / 매출채권", "trade_receivables"),
        ("Inventory / 재고", "inventory"),
        ("Current Assets / 유동자산", "current_assets"),
        ("Total Assets / 자산총계", "total_assets"),
        ("Current Liabilities / 유동부채", "current_liabilities"),
        ("Long-term Debt / 장기차입금", "long_term_debt"),
        ("Total Liabilities / 부채총계", "total_liabilities"),
        ("Total Equity / 자본총계", "total_equity"),
    ]
    for i, (label, key) in enumerate(bs_items, start=2):
        ws.cell(row=i, column=1, value=label).font = section_font
        values = bs_data.get(key, [])
        for j, v in enumerate(values, start=2):
            cell = ws.cell(row=i, column=j, value=v)
            cell.font = body_font
            cell.number_format = "#,##0"
    
    ws.column_dimensions["A"].width = 35
    for j in range(2, len(bs_periods) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 13
    
    # ============================================================
    # Tab 5: Cash Flow
    # ============================================================
    ws = wb.create_sheet("Cash Flow")
    cf_data = context.get("cash_flow", {})
    cf_periods = cf_data.get("periods", periods)
    
    ws.cell(row=1, column=1, value="Cash Flow Statement").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    for j, p in enumerate(cf_periods, start=2):
        cell = ws.cell(row=1, column=j, value=p)
        cell.font = header_font
        cell.fill = header_fill
    
    cf_items = [
        ("Operating Cash Flow / 영업활동현금흐름", "operating_cash_flow"),
        ("CapEx / 유형자산취득", "capex"),
        ("Free Cash Flow", "free_cash_flow"),
        ("Dividends Paid / 배당금", "dividends_paid"),
        ("Share Repurchases / 자사주매입", "share_repurchases"),
    ]
    for i, (label, key) in enumerate(cf_items, start=2):
        ws.cell(row=i, column=1, value=label).font = section_font
        values = cf_data.get(key, [])
        for j, v in enumerate(values, start=2):
            cell = ws.cell(row=i, column=j, value=v)
            cell.font = body_font
            cell.number_format = "#,##0"
    
    ws.column_dimensions["A"].width = 35
    for j in range(2, len(cf_periods) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 13
    
    # ============================================================
    # Tab 6: DCF
    # ============================================================
    ws = wb.create_sheet("DCF")
    dcf = context.get("dcf", {})
    
    ws.cell(row=1, column=1, value="DCF Valuation").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    
    dcf_rows = [
        ("Enterprise Value", dcf.get("enterprise_value")),
        ("(-) Net Debt", dcf.get("net_debt")),
        ("Equity Value", dcf.get("equity_value")),
        ("÷ Diluted Shares", dcf.get("diluted_shares")),
        ("= Implied Share Price", dcf.get("implied_share_price")),
        ("Current Share Price", dcf.get("current_price")),
        ("Upside / Downside", dcf.get("upside_pct")),
    ]
    for i, (label, v) in enumerate(dcf_rows, start=3):
        ws.cell(row=i, column=1, value=label).font = section_font
        cell = ws.cell(row=i, column=2, value=v)
        cell.font = body_font
        if v is not None:
            if "Upside" in label or "%" in label:
                cell.number_format = "0.0%"
            elif "Price" in label:
                cell.number_format = "#,##0.00"
            else:
                cell.number_format = "#,##0"
    
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    
    # 민감도 표
    sensitivity = dcf.get("sensitivity", {})
    if sensitivity:
        wacc_row = sensitivity.get("wacc", [])
        growth_row = sensitivity.get("growth", [])
        matrix = sensitivity.get("matrix", [])
        
        start_row = 13
        ws.cell(row=start_row, column=1, value="Sensitivity (WACC × Terminal Growth)").font = section_font
        for j, g in enumerate(growth_row, start=2):
            cell = ws.cell(row=start_row + 1, column=j, value=g)
            cell.number_format = "0.0%"
            cell.font = section_font
        for i, w in enumerate(wacc_row, start=start_row + 2):
            cell = ws.cell(row=i, column=1, value=w)
            cell.number_format = "0.0%"
            cell.font = section_font
            for j, val in enumerate(matrix[i - start_row - 2], start=2) if i - start_row - 2 < len(matrix) else []:
                cell = ws.cell(row=i, column=j, value=val)
                cell.number_format = "#,##0"
                cell.font = body_font
    
    # ============================================================
    # Save
    # ============================================================
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    return True


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--context", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()
    
    with open(args.context, encoding="utf-8") as f:
        context = json.load(f)
    
    if build_excel(context, args.output):
        print(json.dumps({"output": args.output, "status": "success"}, ensure_ascii=False))


if __name__ == "__main__":
    main()
