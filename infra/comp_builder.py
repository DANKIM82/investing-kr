#!/usr/bin/env python3
"""
comp_builder.py - Multi-tab Comparable companies Excel

CLI:
    python infra/comp_builder.py \\
        --context reports/.tmp/comp_context.json \\
        --output reports/{TICKER}_comp_sheet.xlsx
"""

import argparse
import json
import os
import sys


def build_comp_sheet(context, output_path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[error] openpyxl 미설치", file=sys.stderr)
        return False
    
    wb = Workbook()
    
    target = context.get("target", {})
    peers = context.get("peers", [])  # list of dicts with multiples + financials
    
    header_font = Font(name="Malgun Gothic", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
    bold_font = Font(name="Malgun Gothic", size=10, bold=True)
    body_font = Font(name="Malgun Gothic", size=10)
    
    # Tab 1: Cover
    ws = wb.active
    ws.title = "Cover"
    ws["A1"] = f"Comp Sheet — {target.get('name', '')} ({target.get('ticker', '')})"
    ws["A1"].font = Font(name="Malgun Gothic", size=16, bold=True)
    ws.merge_cells("A1:F1")
    ws["A3"] = "Target:"
    ws["B3"] = target.get("name")
    ws["A4"] = "Peers:"
    ws["B4"] = ", ".join(p.get("ticker", "") for p in peers)
    ws["A5"] = "Generated:"
    ws["B5"] = context.get("date", "")
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 60
    
    # Tab 2: Trading Multiples
    ws = wb.create_sheet("Trading Multiples")
    headers = ["Company", "Ticker", "Mkt Cap", "EV", "P/E TTM", "P/E NTM", "EV/EBITDA", "EV/Rev", "P/S", "P/B", "Div Yield"]
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    all_cos = [target] + peers
    for i, co in enumerate(all_cos, start=2):
        is_target = (i == 2)
        font = bold_font if is_target else body_font
        ws.cell(row=i, column=1, value=co.get("name", "")).font = font
        ws.cell(row=i, column=2, value=co.get("ticker", "")).font = font
        ws.cell(row=i, column=3, value=co.get("market_cap")).font = font
        ws.cell(row=i, column=4, value=co.get("enterprise_value")).font = font
        ws.cell(row=i, column=5, value=co.get("pe_ttm")).font = font
        ws.cell(row=i, column=6, value=co.get("pe_forward")).font = font
        ws.cell(row=i, column=7, value=co.get("ev_ebitda")).font = font
        ws.cell(row=i, column=8, value=co.get("ev_revenue")).font = font
        ws.cell(row=i, column=9, value=co.get("ps_ttm")).font = font
        ws.cell(row=i, column=10, value=co.get("pb")).font = font
        ws.cell(row=i, column=11, value=co.get("dividend_yield")).font = font
        
        for j in [3, 4]:
            ws.cell(row=i, column=j).number_format = "#,##0"
        for j in [5, 6, 7, 8, 9, 10]:
            ws.cell(row=i, column=j).number_format = "0.00"
        ws.cell(row=i, column=11).number_format = "0.00%"
    
    # Median row
    median_row = len(all_cos) + 3
    ws.cell(row=median_row, column=1, value="Peer Median").font = bold_font
    
    def _safe_median(vals):
        nums = [v for v in vals if v is not None and isinstance(v, (int, float))]
        if not nums:
            return None
        nums.sort()
        n = len(nums)
        return nums[n // 2] if n % 2 == 1 else (nums[n // 2 - 1] + nums[n // 2]) / 2
    
    multiples_keys = [None, None, "market_cap", "enterprise_value", "pe_ttm", "pe_forward", "ev_ebitda", "ev_revenue", "ps_ttm", "pb", "dividend_yield"]
    for j, k in enumerate(multiples_keys, start=1):
        if k:
            cell = ws.cell(row=median_row, column=j, value=_safe_median([p.get(k) for p in peers]))
            cell.font = bold_font
            if j in [3, 4]:
                cell.number_format = "#,##0"
            elif j == 11:
                cell.number_format = "0.00%"
            else:
                cell.number_format = "0.00"
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    
    # Tab 3: Operating Metrics
    ws = wb.create_sheet("Operating Metrics")
    op_headers = ["Company", "Ticker", "Revenue (TTM)", "Rev Growth (YoY)", "Gross Margin", "Op Margin", "Net Margin", "ROE"]
    for j, h in enumerate(op_headers, start=1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.font = header_font
        cell.fill = header_fill
    
    op_keys = [None, None, "revenue_ttm", "revenue_growth_yoy", "gross_margin", "operating_margin", "profit_margin", "roe"]
    for i, co in enumerate(all_cos, start=2):
        font = bold_font if i == 2 else body_font
        ws.cell(row=i, column=1, value=co.get("name")).font = font
        ws.cell(row=i, column=2, value=co.get("ticker")).font = font
        for j, k in enumerate(op_keys[2:], start=3):
            v = co.get(k)
            cell = ws.cell(row=i, column=j, value=v)
            cell.font = font
            if k == "revenue_ttm":
                cell.number_format = "#,##0"
            else:
                cell.number_format = "0.0%"
    
    for col in range(1, len(op_headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    
    # Tab 4: Notes
    ws = wb.create_sheet("Notes")
    ws["A1"] = "Notes & Methodology"
    ws["A1"].font = bold_font
    notes = [
        "Peer selection rationale:",
        context.get("peer_rationale", "Same industry, similar business model, similar scale."),
        "",
        "Data sources:",
        "- KR: DART (financial statements) + pykrx (multiples)",
        "- US: yfinance + SEC EDGAR",
        "- JP: yfinance with .T suffix",
        "",
        "Limitations:",
        "- NTM multiples from yfinance may be limited or stale",
        "- Cross-currency comparisons not auto-converted",
        "- Audit-grade accuracy not guaranteed",
    ]
    for i, line in enumerate(notes, start=3):
        ws.cell(row=i, column=1, value=line).font = body_font
    ws.column_dimensions["A"].width = 80
    
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    return True


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--context", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()
    
    with open(args.context, encoding="utf-8") as f:
        context = json.load(f)
    
    if build_comp_sheet(context, args.output):
        print(json.dumps({"output": args.output, "status": "success"}, ensure_ascii=False))


if __name__ == "__main__":
    main()
